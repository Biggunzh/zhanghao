#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os
import sys
sys.stdout.reconfigure(encoding='utf-8')

def inspect_xlsx(file_path):
    """使用openpyxl查看xlsx文件"""
    from openpyxl import load_workbook
    
    print(f"\n{'='*60}")
    print(f"📊 文件: {os.path.basename(file_path)}")
    print(f"{'='*60}")
    
    wb = load_workbook(file_path, read_only=True, data_only=True)
    print(f"工作表: {wb.sheetnames}")
    
    # 查看第一个工作表
    ws = wb[wb.sheetnames[0]]
    print(f"\n--- 工作表: {ws.title} ---")
    
    # 打印前15行
    rows = list(ws.iter_rows(values_only=True, max_row=15))
    for i, row in enumerate(rows):
        print(f"行{i+1}: {row[:10]}")  # 只显示前10列
    
    wb.close()

def inspect_xls(file_path):
    """使用xlrd查看xls文件"""
    import xlrd
    
    print(f"\n{'='*60}")
    print(f"📊 文件: {os.path.basename(file_path)}")
    print(f"{'='*60}")
    
    book = xlrd.open_workbook(file_path)
    print(f"工作表: {book.sheet_names()}")
    
    # 查看第一个工作表
    sheet = book.sheet_by_index(0)
    print(f"\n--- 工作表: {sheet.name} ---")
    print(f"总行数: {sheet.nrows}, 总列数: {sheet.ncols}")
    
    # 打印前15行
    for i in range(min(15, sheet.nrows)):
        row = sheet.row_values(i)
        print(f"行{i+1}: {row[:10]}")  # 只显示前10列

if __name__ == '__main__':
    base_dir = r'D:\月报自动化\月报原始数据'
    
    # 检查各个文件
    files = [
        ('2026-03月报资源使用率详情列表.xls', 'xls'),
        ('2026-03工单总量.xlsx', 'xlsx'),
        ('2026-03-堡垒机.xlsx', 'xlsx'),
        ('2026-03vpn审计.xlsx', 'xlsx'),
    ]
    
    for f, ftype in files:
        file_path = os.path.join(base_dir, f)
        if os.path.exists(file_path):
            try:
                if ftype == 'xls':
                    inspect_xls(file_path)
                else:
                    inspect_xlsx(file_path)
            except Exception as e:
                print(f"❌ 读取失败 {f}: {e}")
        else:
            print(f"❌ 文件不存在: {file_path}")
