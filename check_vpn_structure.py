#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""检查VPN审计表结构"""
from openpyxl import load_workbook
import sys
sys.stdout.reconfigure(encoding='utf-8')

VPN_FILE = r'D:\月报自动化\月报原始数据\2026-03vpn审计.xlsx'

wb = load_workbook(VPN_FILE, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]

print("="*70)
print("VPN审计表结构")
print("="*70)

# 读取表头
rows = list(ws.iter_rows(values_only=True, max_row=15))
header_row = None
for i, row in enumerate(rows):
    if any('用户名' in str(cell) for cell in row if cell):
        header_row = i
        print(f"\n表头行 (第{i+1}行):")
        for j, cell in enumerate(row):
            if cell:
                print(f"  列{j}: {cell}")
        break

print("\n前10行数据:")
print("-"*70)
for i, row in enumerate(rows[header_row+1:header_row+11], 1):
    print(f"\n行{i}:")
    for j, cell in enumerate(row):
        if cell:
            print(f"  列{j}: {cell}")

wb.close()
