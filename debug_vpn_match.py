#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""诊断VPN用户名匹配问题"""
import os
import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook

# 1. 读取堡垒机数据，查看用户名格式
FORTRESS_FILE = r'D:\月报自动化\月报原始数据\2026-03-堡垒机.xlsx'
VPN_FILE = r'D:\月报自动化\月报原始数据\2026-03vpn审计.xlsx'

print("="*70)
print("1. 堡垒机数据中的用户名样本")
print("="*70)

wb1 = load_workbook(FORTRESS_FILE, read_only=True, data_only=True)
ws1 = wb1[wb1.sheetnames[0]]

users_from_fortress = set()
rows = list(ws1.iter_rows(values_only=True, max_row=20))
header_row = None
for i, row in enumerate(rows):
    if any('资产名' in str(cell) for cell in row if cell):
        header_row = i
        break

for row in ws1.iter_rows(values_only=True, min_row=header_row+2, max_row=header_row+12):
    if len(row) > 6:
        asset_name = str(row[6]).strip() if row[6] else ''  # 资产名列
        print(f"  资产名: {asset_name}")
        if asset_name and '-' in asset_name:
            user = asset_name.split('-')[0].strip()
            users_from_fortress.add(user)
            print(f"    -> 提取用户名: {user}")

wb1.close()

print(f"\n提取到的用户名: {list(users_from_fortress)[:10]}")

print("\n" + "="*70)
print("2. VPN审计数据中的用户名样本")
print("="*70)

wb2 = load_workbook(VPN_FILE, read_only=True, data_only=True)
ws2 = wb2[wb2.sheetnames[0]]

users_from_vpn = set()
rows = list(ws2.iter_rows(values_only=True, max_row=20))
header_row = None
for i, row in enumerate(rows):
    if any('用户名' in str(cell) for cell in row if cell):
        header_row = i
        break

for row in ws2.iter_rows(values_only=True, min_row=header_row+2, max_row=header_row+12):
    if len(row) > 0:
        username = str(row[0]).strip() if row[0] else ''
        if username:
            users_from_vpn.add(username)
            print(f"  用户名: {username}")

wb2.close()

print(f"\nVPN中的用户名: {list(users_from_vpn)[:10]}")

print("\n" + "="*70)
print("3. 匹配检查")
print("="*70)

matched = users_from_fortress & users_from_vpn
print(f"堡垒机提取的用户数: {len(users_from_fortress)}")
print(f"VPN中的用户数: {len(users_from_vpn)}")
print(f"匹配成功的用户数: {len(matched)}")

if matched:
    print(f"\n匹配的用户: {list(matched)[:10]}")
else:
    print("\n❌ 没有匹配的用户！用户名格式可能不同。")
    print("\n堡垒机用户名示例:")
    for u in list(users_from_fortress)[:5]:
        print(f"  '{u}'")
    print("\nVPN用户名示例:")
    for u in list(users_from_vpn)[:5]:
        print(f"  '{u}'")
